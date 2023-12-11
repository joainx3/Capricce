import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.font import BOLD
from tkcalendar import DateEntry
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import util.generic as utl
from util.bdo import conexion

class MostrarFacturas:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title('Mostrar Facturas')
        self.ventana.geometry('1280x720')  # Ajusta el tamaño según tus necesidades
        self.ventana.config(bg='#fcfcfc')
        self.ventana.resizable(width=0, height=0)
        utl.centrar_ventana(self.ventana, 1280, 720)

        # Widgets para filtrar por fechas
        self.lblFechaInicio = tk.Label(self.ventana, text="Fecha de Inicio:", font=('Times', 12))
        self.lblFechaInicio.grid(row=0, column=1, padx=10, pady=10, sticky="e")

        # Usa DateEntry en lugar de Entry para seleccionar la fecha de inicio
        self.entryFechaInicio = DateEntry(self.ventana, font=('Times', 12), width=12, background='darkblue', foreground='white', borderwidth=2)
        self.entryFechaInicio.grid(row=0, column=2, padx=10, pady=10)

        self.lblFechaFin = tk.Label(self.ventana, text="Fecha de Fin:", font=('Times', 12))
        self.lblFechaFin.grid(row=1, column=1, padx=10, pady=10, sticky="e")

        # Usa DateEntry en lugar de Entry para seleccionar la fecha de fin
        self.entryFechaFin = DateEntry(self.ventana, font=('Times', 12), width=12, background='darkblue', foreground='white', borderwidth=2)
        self.entryFechaFin.grid(row=1, column=2, padx=10, pady=10)

        btnFiltrarFecha = tk.Button(self.ventana, text="Filtrar por Fecha", font=('Times', 12, BOLD), command=self.filtrar_por_fecha)
        btnFiltrarFecha.grid(row=2, column=2, padx=10, pady=10)

        # Botón de menú
        btnMenu = tk.Button(self.ventana, text="Menú", font=('Times', 15, BOLD), command=self.abrir_menu)
        btnMenu.grid(row=0, column=0, padx=10, pady=10, sticky="nw")

        # Crear la tabla
        self.crear_tabla()

        # Botón para cambiar estado de despacho
        btnCambiarDespacho = tk.Button(self.ventana, text="Cambiar Estado de Despacho", font=('Times', 15, BOLD), command=self.ver_detalles_factura)
        btnCambiarDespacho.grid(row=1, column=0, padx=10, pady=10, sticky="nw")


        self.ventana.mainloop()

    def abrir_menu(self):
        from forms.form_menu import Menu
        self.ventana.destroy()
        Menu()

    def crear_tabla(self):
        frame_tabla = tk.Frame(self.ventana, bd=0, relief=tk.SOLID, bg='#fcfcfc')
        frame_tabla.grid(row=3, column=0, sticky="nsew")

        # Configuración de la tabla
        columns = ["ID_ORDEN", "Nombre Cliente", "Direccion", "Producto", "Precio", "FechaCompra", "EstadoFactura", "EstadoDespacho", "IVA", "TotalConIVA"]
        self.tree = ttk.Treeview(frame_tabla, columns=columns, show="headings", selectmode="browse")

        # Configuración de las columnas
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)  # Ajusta el ancho de las columnas según tus necesidades

        # Llena la tabla con los datos del procedimiento
        self.llenar_tabla_facturas()

        # Enlaza la selección de la tabla a la función de manejo de eventos
        self.tree.bind("<<TreeviewSelect>>", self.seleccionar_fila)

        # Empaqueta la tabla en el frame
        self.tree.pack(expand=tk.YES, fill=tk.BOTH)

        # Configura el comportamiento de redimensionamiento de la ventana principal
        self.ventana.grid_rowconfigure(3, weight=1)
        self.ventana.grid_columnconfigure(0, weight=1)

    def llenar_tabla_facturas(self):
        try:
            # Utiliza la conexión ya disponible en util.bdo
            with conexion.cursor() as cursor:
                # Ejecuta tu procedimiento almacenado aquí
                cursor.execute("EXEC sp_MostrarFacturasClientes")

                # Obtiene los resultados
                resultados = cursor.fetchall()

                # Llena la tabla con los resultados
                for row in resultados:
                    # Ajusta el orden de los valores según la secuencia de columnas esperada
                    # Concatena los componentes del nombre del cliente
                    nombre_cliente = f"{row[1]} "
                    self.tree.insert("", "end", values=(row[0], nombre_cliente, row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]))

        except Exception as e:
            messagebox.showerror("Error", f"Error al conectar a la base de datos: {e}")

    def seleccionar_fila(self, event):
        # Obtiene la fila seleccionada
        selected_item = self.tree.selection()[0]
        # Puedes acceder a los valores de la fila seleccionada usando item
        values = self.tree.item(selected_item, "values")
        print("Fila seleccionada:", values)

    def ver_detalles_factura(self):
        # Implementa la lógica para cambiar el estado del despacho de la factura seleccionada
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Selecciona una fila antes de cambiar el estado de despacho.")
            return

        # Obtiene el ID_ORDEN de la fila seleccionada
        id_orden = self.tree.item(selected_item, "values")[0]

        # Llama al procedimiento almacenado para cambiar el estado del despacho de la factura
        if self.cambiar_estado_despacho(id_orden):
            messagebox.showinfo("Información", "El estado del despacho ha sido actualizado correctamente.")
            # Programa la recarga de la tabla después de un cierto tiempo (por ejemplo, 500 ms)
            self.ventana.after(500, self.recargar_tabla_facturas)
        else:
            messagebox.showerror("Error", "Error al cambiar el estado del despacho de la factura.")

    def cambiar_estado_despacho(self, id_orden):
        try:
            with conexion.cursor() as cursor:
                # Ejemplo de actualización, ajusta según tu esquema de base de datos
                estado_despacho = 'Producto despachado a cliente'
                cursor.execute("UPDATE dbo.Facturas_clientes SET EstadoDespacho = ? WHERE ID_ORDEN = ?", (estado_despacho, id_orden))
                cursor.execute("UPDATE dbo.ordenCompra SET despacho = ? WHERE ID_ORDEN = ?", (estado_despacho, id_orden))
                conexion.commit()
                return True
        except Exception as e:
            messagebox.showerror("Error", f"Error al cambiar el estado del despacho: {e}")
            return False

    def recargar_tabla_facturas(self):
        # Limpia la tabla actual
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Vuelve a llenar la tabla con los datos actualizados
        self.llenar_tabla_facturas()

    def exportar_a_excel(self):
        # Verificar si hay una fila seleccionada
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Selecciona una fila para exportar a Excel.")
            return

        # Obtener los valores de la fila seleccionada
        values = self.tree.item(selected_item, "values")

        # Crear un cuadro de diálogo para seleccionar la ubicación del archivo Excel
        excel_filename = self.guardar_excel()
        if not excel_filename:
            return

        try:
            # Crear un nuevo libro de trabajo y seleccionar la hoja activa
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Agregar encabezados
            headers = ["ID_ORDEN", "Nombre Cliente", "Direccion", "Producto", "Precio", "FechaCompra", "EstadoFactura", "EstadoDespacho", "IVA", "TotalConIVA"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)

            # Agregar datos
            for col_num, value in enumerate(values, 1):
                sheet.cell(row=2, column=col_num, value=value)

            # Guardar el archivo Excel
            workbook.save(excel_filename)

            messagebox.showinfo("Información", f"Exportación a Excel exitosa. Archivo guardado en {excel_filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {e}")

    def guardar_excel(self):
        # Crear un cuadro de diálogo para seleccionar la ubicación del archivo Excel
        file_dialog = filedialog.asksaveasfile(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        # Obtener la ubicación seleccionada
        if file_dialog:
            excel_filename = file_dialog.name
            file_dialog.close()  # Cerrar el archivo abierto por asksaveasfile
            return excel_filename
        else:
            return None

    def filtrar_por_fecha(self):
        fecha_inicio_str = self.entryFechaInicio.get()
        fecha_fin_str = self.entryFechaFin.get()

        try:
            # Convertir las fechas a objetos datetime con el formato "MM/DD/YY"
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%m/%d/%y")
            fecha_fin = datetime.strptime(fecha_fin_str, "%m/%d/%y")

            # Validar que la fecha de inicio sea anterior o igual a la fecha de fin
            if fecha_inicio > fecha_fin:
                messagebox.showwarning("Advertencia", "La fecha de inicio debe ser anterior o igual a la fecha de fin.")
                return

            # Utilizar la conexión ya disponible en util.bdo
            with conexion.cursor() as cursor:
                # Modificar la consulta para incluir las fechas en la cláusula WHERE
                cursor.execute("EXEC sp_MostrarFacturasClientesConFiltroFecha ?, ?", (fecha_inicio, fecha_fin))

                # Obtener los resultados
                resultados = cursor.fetchall()

                # Limpiar la tabla actual
                for item in self.tree.get_children():
                    self.tree.delete(item)

                # Llenar la tabla con los resultados filtrados por fechas
                for row in resultados:
                    nombre_cliente = f"{row[1]} "
                    self.tree.insert("", "end", values=(row[0], nombre_cliente, row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]))

        except Exception as e:
            messagebox.showerror("Error", f"Error al filtrar por fechas: {e}")
   
'''
    def exportar_a_pdf(self):
        # Verificar si hay una fila seleccionada
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Selecciona una fila para exportar a PDF.")
            return

        # Obtener los valores de la fila seleccionada
        values = self.tree.item(selected_item, "values")

        # Crear un cuadro de diálogo para seleccionar la ubicación del archivo PDF
        pdf_filename = self.guardar_pdf()
        if not pdf_filename:
            return

        try:
            with canvas.Canvas(pdf_filename, pagesize=letter) as pdf:
                # Escribir el contenido en el PDF
                pdf.drawString(100, 800, f"ID_ORDEN: {values[0]}".encode('utf-8'))
                pdf.drawString(100, 780, f"Nombre Cliente: {values[1]}".encode('utf-8'))
                pdf.drawString(100, 760, f"Direccion: {values[2]}".encode('utf-8'))
                # ... agregar más campos según tus necesidades

            messagebox.showinfo("Información", f"Exportación a PDF exitosa. Archivo guardado en {pdf_filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a PDF: {e}")

    def guardar_pdf(self):
        # Crear un cuadro de diálogo para seleccionar la ubicación del archivo PDF
        file_dialog = filedialog.asksaveasfile(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])

        # Obtener la ubicación seleccionada
        if file_dialog:
            pdf_filename = file_dialog.name
            file_dialog.close()  # Cerrar el archivo abierto por asksaveasfile
            return pdf_filename
        else:
            return None
'''

if __name__ == "__main__":
    app = MostrarFacturas()
